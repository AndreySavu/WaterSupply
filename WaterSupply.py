#from tkinter import *
from tkinter import ttk, Tk, END, VERTICAL, BOTH, Menu, filedialog
import TKmv
import sqlite3
import NormGraph as NG
import calculations
import math
import pandas as pd
import openpyxl

class App:

    def create_db(self, name):
        cx = sqlite3.connect(name)
        cu = cx.cursor()
        cu.execute("CREATE TABLE sources (\
	                             name VARCHAR(50) PRIMARY KEY,\
	                             coord1 DOUBLE,\
	                             coord2 DOUBLE,\
                                 H DOUBLE,\
                                 G DOUBLE,\
                                 Gmax DOUBLE,\
	                             P DOUBLE);")
        cu.execute("CREATE TABLE towers (\
	                             name VARCHAR(50) PRIMARY KEY,\
	                             coord1 DOUBLE,\
	                             coord2 DOUBLE,\
                                 H DOUBLE,\
                                 G DOUBLE,\
                                 Hwater DOUBLE,\
                                 V DOUBLE,\
	                             P DOUBLE);")
        cu.execute("CREATE TABLE reservoirs (\
	                             name VARCHAR(50) PRIMARY KEY,\
	                             coord1 DOUBLE,\
	                             coord2 DOUBLE,\
                                 H DOUBLE,\
                                 G DOUBLE,\
                                 Hwater DOUBLE,\
                                 P DOUBLE);")
        cu.execute("CREATE TABLE connectors (\
	                             name VARCHAR(50) PRIMARY KEY,\
	                             coord1 DOUBLE,\
	                             coord2 DOUBLE,\
                                 H DOUBLE,\
                                 G DOUBLE,\
                                 P DOUBLE,\
                                 Pressure DOUBLE);")
        cu.execute("CREATE TABLE consumers (\
	                             name VARCHAR(50) PRIMARY KEY,\
	                             coord1 DOUBLE,\
	                             coord2 DOUBLE,\
                                 H DOUBLE,\
                                 G DOUBLE,\
                                 Pmin DOUBLE,\
                                 P DOUBLE);")
        cu.execute("CREATE TABLE pipes (\
                                 name VARCHAR(50) PRIMARY KEY,\
                                 name1 VARCHAR(50), name2 VARCHAR(50),\
                                 H1 DOUBLE, H2 DOUBLE,\
                                 D DOUBLE, Len DOUBLE, G DOUBLE, R DOUBLE,\
                                 Glost DOUBLE,\
                                 Speed DOUBLE,\
                                 Lambda DOUBLE,\
                                 Lost DOUBLE,\
                                 AcceptPres DOUBLE,\
                                 Material VARCHAR(20),\
                                 hasValve INTEGER,\
                                 opened INTEGER);")
        cu.execute("CREATE TABLE map (\
	                id INTEGER DEFAULT 1,\
	                pos1 DOUBLE,\
	                pos2 DOUBLE,\
	                zoom INTEGER);")
        cx.commit()
        cx.close()

    def clear_db(self, name):
        cx = sqlite3.connect(name)
        cu = cx.cursor()
        cu.execute("DELETE FROM sources;")
        cu.execute("DELETE FROM towers;")
        cu.execute("DELETE FROM reservoirs;")
        cu.execute("DELETE FROM connectors;")
        cu.execute("DELETE FROM consumers;")
        cu.execute("DELETE FROM pipes;")
        cu.execute("DELETE FROM map;")
        cx.commit()
        cx.close()

    def clear_xlsx(self, name):
        list = ['Sources', 'WaterTowers', 'CounterReservoirs', 'Connectors', 'Consumers', 'Pipes']
        try:
            wb = openpyxl.load_workbook(name)
            for sheet in list:
                pfd = wb[sheet]
                wb.remove(pfd)
            wb.save(name)
        except:
            new_wb = openpyxl.Workbook()
            for sheet in list:
                new_wb.create_sheet(sheet)
            new_wb.save(name)

    def add_marker_source(self, coords):
        adress = str(TKmv.convert_coordinates_to_address(coords[0], coords[1]).street) + \
            " " + str(TKmv.convert_coordinates_to_address(coords[0], coords[1]).housenumber)
        dat = "Source" + str(self.last_N_of_object[0] + 1)
        self.map_widget.set_marker(coords[0], coords[1],
                                    text = "ИСТ " + adress,
                                    icon = "icons/source.png",
                                    data = dat)                          
        self.gr.add_vertex(dat, coords[0], coords[1], (0, 0, 0, 0))
        self.last_N_of_object[0] += 1

    def add_marker_tower(self, coords):
        adress = str(TKmv.convert_coordinates_to_address(coords[0], coords[1]).street) + \
            " " + str(TKmv.convert_coordinates_to_address(coords[0], coords[1]).housenumber)
        dat = "WaterTower" + str(self.last_N_of_object[1] + 1)
        self.map_widget.set_marker(coords[0], coords[1],
                                                text = "ВНБ " + adress,
                                                icon = "icons/tower.png",
                                                data = dat)                          
        self.gr.add_vertex(dat, coords[0], coords[1], (0, 0, 0, 0, 0))
        self.last_N_of_object[1] += 1
        
    def add_marker_reservoir(self, coords):
        adress = str(TKmv.convert_coordinates_to_address(coords[0], coords[1]).street) + \
            " " + str(TKmv.convert_coordinates_to_address(coords[0], coords[1]).housenumber)
        dat = "CounterReservoir" + str(self.last_N_of_object[2] + 1)
        self.map_widget.set_marker(coords[0], coords[1],
                                    text = "КР " + adress,
                                    icon = "icons/reservoir.png",
                                    data = dat)                          
        self.gr.add_vertex(dat, coords[0], coords[1], (0, 0, 0, 0))
        self.last_N_of_object[2] += 1

    def add_marker_consumer(self, coords):
        adress = str(TKmv.convert_coordinates_to_address(coords[0], coords[1]).street) + \
            " " + str(TKmv.convert_coordinates_to_address(coords[0], coords[1]).housenumber)
        dat = "Consumer" + str(self.last_N_of_object[4] + 1)
        self.map_widget.set_marker(coords[0], coords[1],
                                    text = "ПТР " + adress,
                                    icon = "icons/consumer.png",
                                    data = dat)
        self.gr.add_vertex(dat, coords[0], coords[1], (0, 0, 0, 0))
        self.last_N_of_object[4] += 1

    def add_marker_connector(self, coords):
        dat = "Connector" + str(self.last_N_of_object[3] + 1)
        self.map_widget.set_marker(coords[0], coords[1],
                                    text = "РЗВ " + str(round(coords[0],4)) + "; " + str(round(coords[1],4)),
                                    icon = "icons/connector.png",
                                    data = dat)
        self.gr.add_vertex(dat, coords[0], coords[1], (0, 0, 0, 0))
        self.last_N_of_object[3] += 1

    def add_line(self, coords):
        getfirstpoint = False
        getsecondpoint = False
        if self.firstpoint is None:
            for marker in self.map_widget.canvas_marker_list:
                if not (marker.get_canvas_pos(coords)[0] > marker.get_canvas_pos(marker.position)[0] + 20 or
                    marker.get_canvas_pos(coords)[0] < marker.get_canvas_pos(marker.position)[0] - 20 or
                    marker.get_canvas_pos(coords)[1] > marker.get_canvas_pos(marker.position)[1] + 20 or
                    marker.get_canvas_pos(coords)[1] < marker.get_canvas_pos(marker.position)[1] - 20):
                        self.firstpoint = marker
                        getfirstpoint = True
                        break
            if not getfirstpoint:
                print("Первая точка не найдена")
        else:
            for marker in self.map_widget.canvas_marker_list:
                if  not (marker.get_canvas_pos(coords)[0] > marker.get_canvas_pos(marker.position)[0] + 20 or
                        marker.get_canvas_pos(coords)[0] < marker.get_canvas_pos(marker.position)[0] - 20 or
                        marker.get_canvas_pos(coords)[1] > marker.get_canvas_pos(marker.position)[1] + 20 or
                        marker.get_canvas_pos(coords)[1] < marker.get_canvas_pos(marker.position)[1] - 20) and\
                        self.firstpoint.data != marker.data:
                    self.secondpoint = marker
                    getsecondpoint = True
                    buff = [self.firstpoint.position, self.secondpoint.position]
                    
                    current_amount = len(self.gr.get_all_edges())
                    dat = "Edge" + str(self.last_N_of_object[5] + 1)
                    self.gr.add_edge(dat,
                                        self.firstpoint.data,
                                        self.secondpoint.data, (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 'стальные неновые', 0, 0))
                    
                    if current_amount != len(self.gr.get_all_edges()):
                        self.map_widget.set_polygon(buff, data=dat)
                        self.last_N_of_object[5] += 1

                    self.firstpoint = None
                    self.secondpoint = None
                    break
            if not getsecondpoint:
                print("Вторая точка не найдена")


    def delete_marker(self, coords):
        buff = []
        for marker in self.map_widget.canvas_marker_list:
            if not (marker.get_canvas_pos(coords)[0] > marker.get_canvas_pos(marker.position)[0] + 20 or
                marker.get_canvas_pos(coords)[0] < marker.get_canvas_pos(marker.position)[0] - 20 or
                marker.get_canvas_pos(coords)[1] > marker.get_canvas_pos(marker.position)[1] + 20 or
                marker.get_canvas_pos(coords)[1] < marker.get_canvas_pos(marker.position)[1] - 20):   
                    
                for poly in self.map_widget.canvas_polygon_list:
                    if self.gr.get_edge(poly.data)[1] == marker.data or self.gr.get_edge(poly.data)[2] == marker.data:
                        buff.append(poly)
                self.gr.remove_vertex(marker.data)
                self.map_widget.delete(marker)
                for poly in buff:
                    self.map_widget.delete(poly)
                break

    def delete_line(self, coords):
        for poly in self.map_widget.canvas_polygon_list:
            a = poly.get_canvas_pos(poly.position_list[0], 
                                    self.root.winfo_width()*0.7, 
                                    self.root.winfo_height())
            b = poly.get_canvas_pos(poly.position_list[1], 
                                    self.root.winfo_width()*0.7, 
                                    self.root.winfo_height())
            c = poly.get_canvas_pos(coords, 
                                    self.root.winfo_width()*0.7, 
                                    self.root.winfo_height())
            if abs(math.sqrt((a[0] - b[0])**2 + (a[1] - b[1])**2) - 
                   (math.sqrt((a[0] - c[0])**2 + (a[1] - c[1])**2) 
                    + math.sqrt((b[0] - c[0])**2 + (b[1] - c[1])**2))) < 0.4:
                   self.gr.remove_edge(poly.data)
                   self.map_widget.delete(poly)

    def properties(self, _name):
        width=self.root.winfo_width()*0.7
        
        self.close_properties()
        _object = self.gr.get_vertex(_name)
        match _name[:4]:
            case 'Sour': #источник водоснабжения
                _type = 'Источник'
                list_props_name = self.list_props_source
            case 'Wate': #водонапорная башня
                _type = 'Водонапорная башня'
                list_props_name = self.list_props_watertower
            case 'Coun': #контррезервуар
                _type = 'Контррезервуар'
                list_props_name = self.list_props_counterreservior
            case 'Conn':
                _type = 'Узер(разветвление)'
                list_props_name = self.list_props_connector
            case 'Cons':
                _type = 'Потребитель'
                list_props_name = self.list_props_consumer

        self.type.config(text = _type)
        self.name.config(text = _name)
        self.coordinates.config(text = 'Координаты: ' + str(_object[1]) + " -- " + str(_object[2]))
        self.type.place( x=width+15, y=30)
        self.name.place( x=width+15, y=50)
        self.coordinates.place( x=width+15, y=70)
        
        for i in range (len(list_props_name)):          
            self.props_name[i].config(text= list_props_name[i])
            self.props_name[i].place(x=width+15, y=120+i*20)
            self.props_value[i].insert(0, str(_object[3][i]))
            self.props_value[i].place(x=width+200, y=120+i*20) 

        self.save_button.config(command= self.save_properties)
        self.cancel_button.place(x= self.root.winfo_width()-300, y= self.root.winfo_height()-50)
        self.save_button.place(x= self.root.winfo_width()-200, y= self.root.winfo_height()-50)
    
    def properties_line(self, _name):
        width=self.root.winfo_width()*0.7
        
        self.close_properties()
        _object = self.gr.get_edge(_name)
        
        _type = 'Участок'
        list_props_name = self.list_props_pipe
        
        self.type.config(text = _type)
        self.name.config(text = _name)
        self.coordinates.config(text = 'Конечные точки: ' + str(_object[1]) + " -- " + str(_object[2]))
        self.type.place( x=width+15, y=30)
        self.name.place( x=width+15, y=50)
        self.coordinates.place( x=width+15, y=70)
        
        for i in range (len(list_props_name)):          
            self.props_name[i].config(text= list_props_name[i]) 
            self.props_value[i].insert(0, str(_object[3][i]))
            self.props_name[i].place( x=width+15, y=120+i*20)  
            self.props_value[i].place( x=width+200, y=120+i*20) 
        last_posy = 120+len(list_props_name)*20
        
        self.props_name[11].config(text= 'Материал трубы')
        self.props_name[11].place( x=width+15, y=last_posy+20) 
        self.material.current(self.material_list.index(_object[3][11]))
        self.material.place(x=width+200, y=last_posy+20) 
        self.props_name[12].config(text= 'Наличие запорной арматуры')
        self.props_name[12].place( x=width+15, y=last_posy+40)
        self.has_valve.current(_object[3][12]) 
        self.has_valve.place(x=width+200, y=last_posy+40)
        self.props_name[13].config(text= 'Процент открытия')
        self.props_name[13].place( x=width+15, y=last_posy+60) 
        self.opened_valve.insert(0, str(_object[3][13]))
        self.opened_valve.place(x=width+200, y=last_posy+60)

        self.save_button.config(command= self.save_properties_line)
        self.cancel_button.place(x= self.root.winfo_width()-300, y= self.root.winfo_height()-50)
        self.save_button.place(x= self.root.winfo_width()-200, y= self.root.winfo_height()-50)

    def save_properties(self):
        name = self.name.cget("text")
        coords = self.coordinates.cget("text")[12:].split(' -- ')
        value = []
        match name[:4]:
            case 'Sour': 
                n_values = 4
            case 'Wate': 
                n_values = 5
            case 'Coun': 
                n_values = 4
            case 'Conn': 
                n_values = 4
            case 'Cons': 
                n_values = 4

        for i in range(n_values):
            value.append(float(self.props_value[i].get()))
        self.gr.update_vertex(name,float(coords[0]),float(coords[1]),tuple(value))

        print(self.gr.get_all_vertexes())
    
    def save_properties_line(self):
        name = self.name.cget("text")
        points = self.coordinates.cget("text")[16:].split(' -- ')
        value = []
        for i in range (11):
            value.append(float(self.props_value[i].get()))
        if self.has_valve.get() == 'нет':
            state = 0
        else:
            state = 1
        value.append(self.material.get())
        value.append(state)
        value.append(float(self.opened_valve.get()))
        self.gr.update_edge(name,points[0],points[1],tuple(value))

        print(self.gr.get_all_edges())

    def close_properties(self):
        self.type.place_forget()
        self.name.place_forget()
        self.coordinates.place_forget()
        for item in self.props_name:
            item.place_forget()
        for item in self.props_value:
            item.place_forget()
            item.delete(0,END)
        self.save_button.place_forget()
        self.cancel_button.place_forget()
        self.has_valve.place_forget() 
        self.material.place_forget() 
        self.opened_valve.place_forget() 
        self.opened_valve.delete(0,END)   
        
    def open_new_map(self, pos=(55.010159, 82.925716), zoom = 10, path_state = 0):
        self.gr.clear_graph()
        self.close_properties()
        if not path_state:
            self.path_to_file = ''
        self.last_N_of_object = [0, 0, 0, 0, 0, 0]
        self.has_valve.current(0)
        self.material.current(0)
        self.map_widget = TKmv.TkinterMapView(self.root, width=self.root.winfo_width()*0.7,
                                               height=self.root.winfo_height(), corner_radius=0)
        self.map_widget.set_position(pos[0], pos[1])
        self.map_widget.set_zoom(zoom)
        self.map_widget.grid(row=0, column=0)
        #print(TKmv.decimal_to_osm(pos[0], pos[1], zoom))

        self.map_widget.add_right_click_menu_command(label="Добавить источник", 
                                command=self.add_marker_source, 
                                pass_coords=True,
                                )
        self.map_widget.add_right_click_menu_command(label="Добавить водонапорную башню", 
                                command=self.add_marker_tower, 
                                pass_coords=True,
                                )
        self.map_widget.add_right_click_menu_command(label="Добавить контеррезервуар", 
                                command=self.add_marker_reservoir, 
                                pass_coords=True,
                                )        
        self.map_widget.add_right_click_menu_command(label="Добавить потребителя", 
                                command=self.add_marker_consumer, 
                                pass_coords=True,
                                )
        self.map_widget.add_right_click_menu_command(label="Добавить узел", 
                                command=self.add_marker_connector, 
                                pass_coords=True,
                                )
        self.map_widget.add_right_click_menu_command_markers(label="Добавить соединение", 
                                command=self.add_line,
                                pass_coords=True 
                                )
        self.map_widget.add_right_click_menu_command_markers(label="Удалить объект", 
                                command=self.delete_marker,
                                pass_coords=True 
                                )
        self.map_widget.add_right_click_menu_command_markers(label="Свойства", 
                                command=self.properties,
                                pass_coords=False 
                                )
        self.map_widget.add_right_click_menu_command_lines(label="Удалить участок", 
                                command=self.delete_line,
                                pass_coords=True 
                                )
        self.map_widget.add_right_click_menu_command_lines(label="Свойства", 
                                command=self.properties_line,
                                pass_coords=False 
                                )
        
        self.map_widget.save_state_to_file
    
    def load(self):
        self.last_N_of_object = [0, 0, 0, 0, 0, 0]
        self.path_to_file = filedialog.askopenfilename(defaultextension=".txt",
                    filetypes=[("database",".db"), ("Excel",".xlsx")],
                    initialdir="./schemas")
        
        if self.path_to_file[-3:] == '.db':
            cx = sqlite3.connect(self.path_to_file)
            cu = cx.cursor()
            cu.execute("SELECT * FROM map;")
            out0 = cu.fetchall()
            cu.execute("SELECT * FROM sources;")
            out1 = cu.fetchall()
            cu.execute("SELECT * FROM towers;")
            out2 = cu.fetchall()
            cu.execute("SELECT * FROM reservoirs;")
            out3 = cu.fetchall()
            cu.execute("SELECT * FROM connectors;")
            out4 = cu.fetchall()
            cu.execute("SELECT * FROM consumers;")
            out5 = cu.fetchall()
            cu.execute("SELECT * FROM pipes;")
            out6 = cu.fetchall()
            cx.commit()
            cx.close()
            self.open_new_map((out0[0][1], out0[0][2]), out0[0][3], 1)#инициализация карты
        
        if self.path_to_file[-5:] == '.xlsx':
            out0 = pd.read_excel(self.path_to_file, sheet_name='Map').values.tolist()
            out1 = pd.read_excel(self.path_to_file, sheet_name='Sources').values.tolist()
            out2 = pd.read_excel(self.path_to_file, sheet_name='WaterTowers').values.tolist()
            out3 = pd.read_excel(self.path_to_file, sheet_name='CounterReservoirs').values.tolist()
            out4 = pd.read_excel(self.path_to_file, sheet_name='Connectors').values.tolist()
            out5 = pd.read_excel(self.path_to_file, sheet_name='Consumers').values.tolist()
            out6 = pd.read_excel(self.path_to_file, sheet_name='Pipes').values.tolist()
            self.open_new_map((out0[0][0], out0[0][1]), out0[0][2], 1)#инициализация карты
 
        for row in out1:#информация об источника
            self.map_widget.set_marker(row[1], row[2], text = "ИСТ " + str(TKmv.convert_coordinates_to_address(row[1], row[2]).street) + 
                                                    " " + str(TKmv.convert_coordinates_to_address(row[1], row[2]).housenumber),
                                                    icon="icons/source.png", data = row[0])
            self.gr.add_vertex(row[0], row[1], row[2], (row[3], row[4], row[5], row[6]))
        if len(out1)!=0:
            self.last_N_of_object[0]=int(row[0][6:])
        for row in out2:#информация о водонапорных башнях
            self.map_widget.set_marker(row[1], row[2], text = "ВНБ " + str(TKmv.convert_coordinates_to_address(row[1], row[2]).street) + 
                                                    " " + str(TKmv.convert_coordinates_to_address(row[1], row[2]).housenumber),
                                                    icon="icons/tower.png", data = row[0])
            self.gr.add_vertex(row[0], row[1], row[2], (row[3], row[4], row[5], row[6], row[7]))
        if len(out2)!=0:    
            self.last_N_of_object[1]=int(row[0][10:])
        for row in out3:#информация о контррезервуарах
            self.map_widget.set_marker(row[1], row[2], text = "КР " + str(TKmv.convert_coordinates_to_address(row[1], row[2]).street) + 
                                                    " " + str(TKmv.convert_coordinates_to_address(row[1], row[2]).housenumber),
                                                    icon="icons/reservoir.png", data = row[0])
            self.gr.add_vertex(row[0], row[1], row[2], (row[3], row[4], row[5], row[6]))
        if len(out3)!=0:
            self.last_N_of_object[2]=int(row[0][16:])
        for row in out4:#информация об узлах
            self.map_widget.set_marker(row[1], row[2], text = "РЗВ " + str(TKmv.convert_coordinates_to_address(row[1], row[2]).street) + 
                                                    " " + str(TKmv.convert_coordinates_to_address(row[1], row[2]).housenumber),
                                                    icon="icons/connector.png", data = row[0])
            self.gr.add_vertex(row[0], row[1], row[2], (row[3], row[4], row[5], row[6]))
        if len(out4)!=0:
            self.last_N_of_object[3]=int(row[0][9:])
        for row in out5:#информация о потребителях
            self.map_widget.set_marker(row[1], row[2], text = "ПТР " + str(TKmv.convert_coordinates_to_address(row[1], row[2]).street) + 
                                                    " " + str(TKmv.convert_coordinates_to_address(row[1], row[2]).housenumber),
                                                    icon="icons/consumer.png", data = row[0])
            self.gr.add_vertex(row[0], row[1], row[2], (row[3], row[4], row[5], row[6]))
        if len(out5)!=0:
            self.last_N_of_object[4]=int(row[0][8:])
        for row in out6:#информация об участках
            point1 = self.gr.get_vertex(row[1])
            point2 = self.gr.get_vertex(row[2])
            #получаем координаты концов участка
            buff = [(point1[1], point1[2]), (point2[1], point2[2])]
            self.map_widget.set_polygon(buff, data=row[0])
            
            self.gr.add_edge(row[0], row[1], row[2], (row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],
                                                      row[11],row[12],row[13],row[14], row[15],row[16]))
        if len(out6)!=0:
            self.last_N_of_object[5]=int(row[0][4:])
        print(self.last_N_of_object)    

    def save(self):
        self.close_properties()
        pos = self.map_widget.convert_canvas_coords_to_decimal_coords(self.map_widget.width / 2, self.map_widget.height / 2)

        if self.path_to_file[-5:] == '.xlsx':
            df0 = pd.DataFrame({item:[] for item in ['Широта', 'Долгота', 'Приближение']})
            df0.loc[ len(df0.index )] = [pos[0], pos[1], self.map_widget.zoom]  

            self.clear_xlsx(self.path_to_file)
            df1 = self.make_dataframe('Source')
            df2 = self.make_dataframe('WaterTower')
            df3 = self.make_dataframe('CounterReservoir')
            df4 = self.make_dataframe('Connector')
            df5 = self.make_dataframe('Consumer')
            df6 = self.make_dataframe('Pipe')
            
            writer = pd.ExcelWriter(self.path_to_file) 
            df0.to_excel(writer, 'Map', index=False)
            df1.to_excel(writer, 'Sources', index=False)
            df2.to_excel(writer, 'WaterTowers', index=False)
            df3.to_excel(writer, 'CounterReservoirs', index=False)
            df4.to_excel(writer, 'Connectors', index=False)
            df5.to_excel(writer, 'Consumers', index=False)
            df6.to_excel(writer, 'Pipes', index=False)
            writer.save()

        elif self.path_to_file[-3:] == '.db':
            try:
                self.clear_db(self.path_to_file)
            except:
                pass
            cx = sqlite3.connect(self.path_to_file)
            cu = cx.cursor()
            
            cu.execute("INSERT OR REPLACE INTO map (pos1, pos2, zoom) VALUES (?, ?, ?);", (pos[0], pos[1], self.map_widget.zoom))

            for item in self.gr.get_all_vertexes():
                if item[0][:4] == "Sour":
                    cu.execute("INSERT OR REPLACE INTO sources (name, coord1, coord2, H, G, Gmax, P) VALUES (?, ?, ?, ?, ?, ?, ?);", 
                            (item[0], item[1], item[2], 
                             item[3][0], item[3][1], item[3][2], item[3][3]))     
                if item[0][:4] == "Wate":
                    cu.execute("INSERT OR REPLACE INTO towers (name, coord1, coord2, H, G, Hwater, V, P) VALUES (?, ?, ?, ?, ?, ?, ?, ?);", 
                            (item[0], item[1], item[2], 
                             item[3][0], item[3][1], item[3][2], item[3][3], item[3][4]))  
                if item[0][:4] == "Coun":
                    cu.execute("INSERT OR REPLACE INTO reservoirs (name, coord1, coord2, H, G, Hwater, P) VALUES (?, ?, ?, ?, ?, ?, ?);", 
                            (item[0], item[1], item[2], 
                             item[3][0], item[3][1], item[3][2], item[3][3]))    
                if item[0][:4] == "Conn":
                      cu.execute("INSERT OR REPLACE INTO connectors (name, coord1, coord2, H, G, P, Pressure) VALUES (?, ?, ?, ?, ?, ?, ?);", 
                            (item[0], item[1], item[2], 
                             item[3][0], item[3][1], item[3][2], item[3][3]))  
                if item[0][:4] == "Cons":
                    cu.execute("INSERT OR REPLACE INTO consumers (name, coord1, coord2, H, G, Pmin, P) VALUES (?, ?, ?, ?, ?, ?, ?);", 
                            (item[0], item[1], item[2], 
                             item[3][0], item[3][1], item[3][2], item[3][3]))  
            for item in self.gr.get_all_edges():
                cu.execute("INSERT OR REPLACE INTO pipes (name, name1, name2, H1, H2, D, Len, G, R, Glost, Speed, Lambda, Lost,\
                        AcceptPres, Material, hasValve, opened) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", 
                        (item[0], item[1], item[2], item[3][0], item[3][1], item[3][2], item[3][3], 
                                                    item[3][4], item[3][5], item[3][6], item[3][7], 
                                                    item[3][8], item[3][9], item[3][10], item[3][11], 
                                                    item[3][12], item[3][13]))
            cx.commit()
            cx.close()
        else:
            self.save_as()    

    def save_as(self):
        self.path_to_file = filedialog.asksaveasfilename(defaultextension=".txt",
                filetypes=[("database",".db"), ("Excel",".xlsx")],
                initialdir="./schemas")
        if self.path_to_file[-3:] == '.db':
            self.create_db(self.path_to_file)
            self.save()
        if self.path_to_file[-5:] == '.xlsx':
            self.save()

    def make_dataframe(self, _type):
        if _type == 'Pipe':
            props_list = ['Название','Начало','Конец'] + self.list_props_pipe + \
                        ['Материал трубы', 'Наличие запорной арматуры', 'Процент открытия']
            
            df = pd.DataFrame({item:[] for item in props_list})
            for item in self.gr.get_all_edges():
                df.loc[ len(df.index )] = [item[0], item[1], item[2]] + [i for i in item[3]]    
            return df
        else:
            props_list = ['Название','Широта','Долгота']
            if _type == 'Source':
                props_list += self.list_props_source     
            if _type == 'WaterTower':
                props_list += self.list_props_watertower    
            if _type == 'CounterReservoir':
                props_list += self.list_props_counterreservior   
            if _type == 'Connector':
                props_list += self.list_props_connector    
            if _type == 'Consumer':
                props_list += self.list_props_consumer  
            _type = _type[:4]   
            df = pd.DataFrame({item:[] for item in props_list}) 
            for item in self.gr.get_all_vertexes():
                if item[0][:4] == _type:
                    df.loc[ len(df.index )] = [item[0], item[1], item[2]] + [i for i in item[3]]
            return df

    def show_group_of_objects(self, _type):
        window = Tk()
        window.title(_type)
        window.geometry("1200x500")
        df = self.make_dataframe(_type)
        
        columns = ['#'+ str(i+1) for i in range (len(df.columns))]
        tree = ttk.Treeview(window, show="headings", columns=columns)
        for i in range (len(columns)):
            tree.heading(columns[i], text=df.columns[i])
            tree.column(columns[i], width=int(1200/len(columns)))

        ysb = ttk.Scrollbar(window, orient= VERTICAL, command=tree.yview)
        tree.configure(yscroll=ysb.set)

        for item in df.values:
            tree.insert("", END, values=tuple(item))
        #self.tree.bind("<>", self.print_selection)
        tree.pack(fill=BOTH, expand=True)

    def make_commute_task(self):
        def on_select():
            if not tree.selection():
                return
            selected_item = tree.selection()[0]
            values = tree.item(selected_item, option="values")
            return values[0]
        
        def calculate():
            result.delete(0,END)
            result.insert(0,calculations.commute_task(self.gr, on_select()))
        
        window = Tk()
        window.title("Коммуникативная задача")
        window.geometry("700x500")
        props_list = ['Название','Широта','Долгота']
        table_name = ttk.Label(window,text='Список вершин')
        msg = ttk.Label(window,text='Выберите вершину для отключения')
        res_text = ttk.Label(window,text='Результат: ')
        result = ttk.Entry(window,width=100)
        start = ttk.Button(window,text='Расчитать', command= calculate)
        columns = ['#'+ str(i+1) for i in range (3)]
        tree = ttk.Treeview(window, show="headings", columns=columns)
        for i in range (len(columns)):
            tree.heading(columns[i], text= props_list[i])
            tree.column(columns[i], width=int(630/len(columns)))

        ysb = ttk.Scrollbar(window, orient= VERTICAL, command=tree.yview)
        tree.configure(yscroll=ysb.set)

        for item in self.gr.get_all_vertexes():
            tree.insert("", END, values=(item[0],item[1],item[2]))
        
        table_name.place(x=320,y=10)
        tree.place(x=30,y=30)
        msg.place(x=285, y=280)
        res_text.place(x=15,y=380)
        result.place(x=80,y=380)
        start.place(x=330,y=410)



    def __init__(self):
        #source, watertower, counterreservoir, consumer, connector, pipe
        self.last_N_of_object = [0, 0, 0, 0, 0, 0]
        self.path_to_file = ''

        self.list_props_source = ['Высота объекта, м', 'Расход воды, м3/час', 
                                   'Максимальный расход, м3/час', 'Напор на выходе, м', ]
        self.list_props_watertower = ['Высота объекта, м', 'Расход воды, м3/час',
                            'Высота воды, м', 'Объем запаса воды, м3', 'Напор, м']
        self.list_props_counterreservior = ['Высота объекта, м', 'Расход воды, м3/час',
                            'Высота воды, м', 'Напор, м']
        self.list_props_connector = ['Высота объекта, м', 'Расход воды, м3/час',
                            'Напор, м', 'Давление воды, м']
        self.list_props_consumer = ['Высота объекта, м', 'Расчетный расход воды, м3/час', 
                            'Минимальный напор, м', 'Напор, м']
        self.list_props_pipe = ['Высота начала, м', 'Высота конца, м', 
                                'Внутренний диаметр, м', 'Длина участка, м',
                                'Расход воды, м3/час', 'Гидравлическое сопротивление, м/(т/ч)2',
                                'Потери напора на участке, м', 'Скорость движения воды, м/с',
                                'Коэффициент гидравл. трения(λ)', 'Утечка, м3/ч', 
                                'Условно допустимое давление, м']
        #граф-----------------------------------
        self.gr = NG.Graph()
        #объект Tkinter--------------
        self.firstpoint = None
        self.secondpoint = None
        self.root = Tk()
        self.root.geometry('1200x600')
        self.root.title("WaterSupply")
        #self.root.state('zoomed')
        #вывод текста справа
        self.type = ttk.Label()
        self.name = ttk.Label()
        self.coordinates = ttk.Label()
        self.props_name = [ttk.Label() for i in range(14)]
        self.props_value = [ttk.Entry(width=25) for i in range(11)] 
        self.save_button = ttk.Button(text='Сохранить')
        self.cancel_button = ttk.Button(text='Отмена', command= self.close_properties)
        self.material_list = ['стальные неновые','стальные новые', 'чугунные неновые', 'чугунные новые',
                               'асбестоцементные', 'полиэтиленовые', 'с цемент.-песч. покрытием']
        self.material = ttk.Combobox(values=self.material_list, state= "readonly", width=22) #материал трубы участка
        self.has_valve_list = ['нет','да']
        self.has_valve = ttk.Combobox(values=self.has_valve_list, state= "readonly", width=15) #наличие запороной арматуры
        self.opened_valve = ttk.Entry(width=15) #процент открытия арматуры    
        #верхний бар------------------------------------------------------
        self.mainmenu = Menu(self.root) 
        self.root.config(menu=self.mainmenu) 
     
        #работа с файлами
        self.filemenu = Menu(self.mainmenu, tearoff=0)
        self.filemenu.add_command(label="Создать новую схему",
                                  command=self.open_new_map)
        self.filemenu.add_command(label="Открыть...",
                                  command=self.load)
        self.filemenu.add_command(label="Сохранить...",
                                  command=self.save)
        self.filemenu.add_command(label="Сохранить как...", 
                                  command=self.save_as)
        self.filemenu.add_command(label="Выход",
                                  command=self.root.destroy)
        #расчеты
        self.calculationsmenu = Menu(self.mainmenu, tearoff=0)
        self.calculationsmenu.add_command(label="Коммуникативная задача",
                                          command=self.make_commute_task)
        self.calculationsmenu.add_command(label="Поверочный расчет")
        #вывод сводных таблиц по типам объектов
        self.tablesmenu = Menu(self.mainmenu, tearoff=0)
        self.tablesmenu.add_command(label="По источникам", 
                                    command=lambda: self.show_group_of_objects('Source'))
        self.tablesmenu.add_command(label="По водонапорным башням", 
                                    command=lambda: self.show_group_of_objects('WaterTower'))
        self.tablesmenu.add_command(label="По контррезервуарам", 
                                    command=lambda: self.show_group_of_objects('CounterReservoir'))
        self.tablesmenu.add_command(label="По узлам (разветвлениям)", 
                                    command=lambda: self.show_group_of_objects('Connector'))
        self.tablesmenu.add_command(label="По потребителям", 
                                    command=lambda: self.show_group_of_objects('Consumer'))
        self.tablesmenu.add_command(label="По участкам", 
                                    command=lambda: self.show_group_of_objects('Pipe'))

        #справка
        self.helpmenu = Menu(self.mainmenu, tearoff=0)
        self.helpmenu.add_command(label="Помощь")
        self.helpmenu.add_command(label="О программе")

        self.mainmenu.add_cascade(label="Файл", menu=self.filemenu)
        self.mainmenu.add_cascade(label="Расчеты", menu=self.calculationsmenu)
        self.mainmenu.add_cascade(label="Сводные таблицы", menu=self.tablesmenu)
        self.mainmenu.add_cascade(label="Справка", menu=self.helpmenu)
        
        self.root.mainloop()



app = App()
